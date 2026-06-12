import base64
import io
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:  # pragma: no cover - handled gracefully at runtime
    openpyxl = None


# Canonical input columns, in the order they are written back into the result
# file. The first element of each tuple is the internal key, the second is the
# header label shown to the user. Headers in the uploaded file are matched to a
# key by stripping + lower-casing (so "   First Name", "City " etc. still map).
INPUT_KEYS = [
    ('first_name', 'First Name'),
    ('last_name', 'Last Name'),
    ('email', 'Email'),
    ('phone', 'Phone'),
    ('person_linkedin', 'LinkedinURL'),
    ('designation', 'Designation'),
    ('company_name', 'Company Name'),
    ('website', 'Website'),
    ('company_linkedin', 'Linkedin'),
    ('industry', 'Industry'),
    ('annual_revenue', 'Annual Revenue'),
    ('employees', 'Employees'),
    ('city', 'City'),
    ('state', 'State'),
    ('country', 'Country'),
]

# Map of normalised header -> internal key (built once from INPUT_KEYS, with a
# couple of common aliases added so slightly different headers still match).
COLUMN_MAP = {label.strip().lower(): key for key, label in INPUT_KEYS}
COLUMN_MAP.update({
    'linkedin url': 'person_linkedin',
    'linkedinurl': 'person_linkedin',
    'company linkedin': 'company_linkedin',
    'job position': 'designation',
    'title': 'designation',
})

ROW_STATUS_LABEL = {'success': 'Success', 'skipped': 'Skipped', 'failed': 'Failed'}


class ContactUploadWizard(models.TransientModel):
    _name = 'contact.upload.wizard'
    _description = 'Upload Contacts & Accounts'

    state = fields.Selection([
        ('draft', 'Upload'),
        ('done', 'Result'),
    ], default='draft')

    upload_file = fields.Binary(string='Excel File')
    file_name = fields.Char(string='File Name')

    result_file = fields.Binary(string='Result File', readonly=True)
    result_file_name = fields.Char(string='Result File Name', readonly=True)

    line_ids = fields.One2many(
        'contact.upload.wizard.line', 'wizard_id', string='Rows', readonly=True)

    # Summary counters (filled in after processing).
    total_rows = fields.Integer(string='Rows Processed', readonly=True)
    account_created_count = fields.Integer(string='Accounts Created', readonly=True)
    account_existing_count = fields.Integer(string='Accounts Matched', readonly=True)
    contact_created_count = fields.Integer(string='Contacts Created', readonly=True)
    contact_skipped_count = fields.Integer(string='Contacts Skipped', readonly=True)
    error_count = fields.Integer(string='Failed Rows', readonly=True)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _clean(value):
        """Return a trimmed string for any cell value. Whole-number floats
        (openpyxl reads numbers as float) are rendered without the trailing
        '.0' so e.g. a phone or revenue cell stays readable."""
        if value is None:
            return ''
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    @staticmethod
    def _to_int(value):
        try:
            return int(float(str(value).replace(',', '').strip()))
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _error_text(exc):
        """Concise, single-line reason for a failed row."""
        message = getattr(exc, 'name', None) or str(exc) or exc.__class__.__name__
        return ' '.join(message.split())[:500]

    def _parse_workbook(self):
        """Decode the upload, return (data_rows, header->column-index map)."""
        if openpyxl is None:
            raise UserError(_(
                "The Python library 'openpyxl' is required to read Excel "
                "files. Please install it on the server (pip install openpyxl)."))
        if not self.upload_file:
            raise UserError(_("Please choose an Excel (.xlsx) file to upload."))
        try:
            data = base64.b64decode(self.upload_file)
            workbook = openpyxl.load_workbook(
                io.BytesIO(data), data_only=True, read_only=True)
        except Exception as exc:
            raise UserError(_(
                "Could not read the file. Please upload a valid .xlsx file.\n\n%s",
                exc))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not rows:
            raise UserError(_("The uploaded file is empty."))

        col_index = {}
        for idx, header in enumerate(rows[0]):
            key = COLUMN_MAP.get(self._clean(header).lower())
            if key and key not in col_index:
                col_index[key] = idx
        if not ({'website', 'company_name', 'email'} & set(col_index)):
            raise UserError(_(
                "Could not find the expected columns. The first row must "
                "contain headers such as 'Company Name', 'Website', 'Email' "
                "and 'First Name'."))
        return rows[1:], col_index

    # ------------------------------------------------------------------
    # Main entry point
    # ------------------------------------------------------------------
    def action_process(self):
        self.ensure_one()
        # Disable mail tracking / auto-subscribe noise while bulk importing.
        Partner = self.env['res.partner'].with_context(
            tracking_disable=True,
            mail_create_nolog=True,
            mail_create_nosubscribe=True,
        )
        data_rows, col_index = self._parse_workbook()

        def cell(row, key):
            idx = col_index.get(key)
            if idx is None or idx >= len(row):
                return ''
            return self._clean(row[idx])

        company_cache = {}   # normalised key -> account (res.partner, is_company=True)
        seen_emails = set()  # emails already handled within THIS file
        results = []         # list of (raw_row, line_dict)

        self.line_ids.unlink()

        for offset, row in enumerate(data_rows):
            excel_row_no = offset + 2  # +1 for header, +1 for 1-based rows

            # Skip fully blank rows silently.
            if not any(self._clean(v) for v in row):
                continue

            website = cell(row, 'website')
            company_name = cell(row, 'company_name')
            first_name = cell(row, 'first_name')
            last_name = cell(row, 'last_name')
            email = cell(row, 'email')
            contact_name = ' '.join(p for p in (first_name, last_name) if p)

            line = {
                'row_number': excel_row_no,
                'company_name': company_name,
                'website': website,
                'contact_name': contact_name,
                'email': email,
                'account_status': '',
                'contact_status': '',
                'row_status': 'success',
                'message': '',
                'company_id': False,
                'partner_id': False,
            }

            # --- Account: match by website, create once, never update -----
            company = Partner.browse()
            try:
                with self.env.cr.savepoint():
                    company = self._resolve_company(
                        Partner, company_cache, website, company_name, row, cell, line)
            except Exception as exc:
                line['row_status'] = 'failed'
                line['account_status'] = 'Error'
                line['message'] = _("Account error: %s", self._error_text(exc))
                _logger.warning("Contact upload row %s (account) failed: %s",
                                excel_row_no, exc)
                results.append((row, line))
                continue  # cannot tag a contact without a resolved account

            # --- Contact: match by email, skip if it already exists -------
            if not (contact_name or email):
                # An account-only row (company with no contact details).
                line['contact_status'] = '—'
                if not line['message']:
                    line['message'] = _("Account-only row (no contact details).")
                results.append((row, line))
                continue

            try:
                with self.env.cr.savepoint():
                    self._resolve_contact(
                        Partner, company, seen_emails, contact_name, email, row, cell, line)
            except Exception as exc:
                line['row_status'] = 'failed'
                line['contact_status'] = 'Error'
                line['message'] = _("Contact error: %s", self._error_text(exc))
                _logger.warning("Contact upload row %s (contact) failed: %s",
                                excel_row_no, exc)

            results.append((row, line))

        # Derive the overall row status for non-failed rows: a row is a
        # "success" only if it actually created something new, otherwise it is
        # a "skipped" (everything already existed).
        for _row, line in results:
            if line['row_status'] == 'failed':
                continue
            created = (line['account_status'] == 'Created'
                       or line['contact_status'] == 'Created')
            line['row_status'] = 'success' if created else 'skipped'

        self._store_results(results, col_index)

        return {
            'type': 'ir.actions.act_window',
            'name': _('Upload Result'),
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    # ------------------------------------------------------------------
    # Account / Contact resolution
    # ------------------------------------------------------------------
    def _resolve_company(self, Partner, cache, website, company_name, row, cell, line):
        """Return the account for this row, creating it at most once.

        Matching priority: website (normalised) first, company name as a
        fallback when no website is given. An existing account is reused as-is
        and never updated."""
        normalized = Partner._normalize_website(website)
        if normalized:
            key = 'web:' + normalized
        elif company_name:
            key = 'name:' + company_name.lower()
        else:
            line['account_status'] = '—'
            return Partner.browse()

        # 1) Already handled earlier in this same file.
        if key in cache:
            company = cache[key]
            line['account_status'] = 'Reused (in file)'
            line['company_id'] = company.id
            return company

        # 2) Already present in the database -> use it, do not update.
        company = Partner.browse()
        if normalized:
            company = Partner._find_company_by_website(website)
        if not company and company_name:
            company = Partner.search([
                ('is_company', '=', True),
                ('name', '=ilike', company_name),
            ], limit=1)
        if company:
            cache[key] = company
            line['account_status'] = 'Already Exists'
            line['company_id'] = company.id
            return company

        # 3) Create a brand new account (once).
        company = Partner.create(self._company_vals(Partner, website, company_name, row, cell))
        cache[key] = company
        line['account_status'] = 'Created'
        line['company_id'] = company.id
        return company

    def _resolve_contact(self, Partner, company, seen_emails, contact_name, email, row, cell, line):
        """Create the contact under ``company`` unless its email already
        exists (in the database or earlier in this file), in which case skip."""
        normalized_email = email.strip().lower() if email else ''

        if normalized_email:
            if normalized_email in seen_emails:
                line['row_status'] = 'skipped'
                line['contact_status'] = 'Skipped - Duplicate Email (in file)'
                line['message'] = _("This email already appears earlier in the file.")
                return
            existing = Partner.search([
                ('email', '=ilike', email),
                ('is_company', '=', False),
            ], limit=1)
            if existing:
                seen_emails.add(normalized_email)
                line['row_status'] = 'skipped'
                line['contact_status'] = 'Skipped - Email Exists'
                line['partner_id'] = existing.id
                line['message'] = _("A contact with email '%s' already exists.", email)
                return

        contact = Partner.create(self._contact_vals(company, contact_name, email, row, cell))
        if normalized_email:
            seen_emails.add(normalized_email)
        line['contact_status'] = 'Created'
        line['partner_id'] = contact.id

    # ------------------------------------------------------------------
    # Value builders
    # ------------------------------------------------------------------
    def _company_vals(self, Partner, website, company_name, row, cell):
        vals = {
            'is_company': True,
            'name': (company_name
                     or Partner._company_name_from_website(website)
                     or website
                     or _('Unknown Account')),
        }
        if website:
            vals['website'] = website
        industry = cell(row, 'industry')
        if industry:
            vals['industry_id'] = self._find_or_create_industry(industry).id
        annual_revenue = cell(row, 'annual_revenue')
        if annual_revenue:
            vals['annual_revenue'] = self._to_int(annual_revenue)
        company_linkedin = cell(row, 'company_linkedin')
        if company_linkedin:
            vals['company_linkedin'] = company_linkedin
        city = cell(row, 'city')
        if city:
            vals['city'] = city
        country = self._find_country(cell(row, 'country'))
        if country:
            vals['country_id'] = country.id
        state = self._find_state(cell(row, 'state'), country)
        if state:
            vals['state_id'] = state.id
        return vals

    def _contact_vals(self, company, contact_name, email, row, cell):
        # NOTE: 'website' is deliberately NOT set on the contact. The
        # res.partner create() override uses a contact's website to (re)resolve
        # or auto-create a parent company; the account is already decided here,
        # so we pass parent_id directly and leave website off to avoid creating
        # a duplicate company.
        vals = {
            'is_company': False,
            'name': contact_name or email or _('Unknown Contact'),
            'parent_id': company.id if company else False,
        }
        if email:
            vals['email'] = email
        phone = cell(row, 'phone')
        if phone:
            vals['phone'] = phone
        designation = cell(row, 'designation')
        if designation:
            vals['function'] = designation
        person_linkedin = cell(row, 'person_linkedin')
        if person_linkedin:
            vals['contact_linkedin'] = person_linkedin
        return vals

    # ------------------------------------------------------------------
    # Lookups for related records
    # ------------------------------------------------------------------
    @api.model
    def _find_or_create_industry(self, name):
        Industry = self.env['res.partner.industry']
        industry = Industry.search([('name', '=ilike', name)], limit=1)
        if not industry:
            industry = Industry.create({'name': name})
        return industry

    @api.model
    def _find_country(self, name):
        Country = self.env['res.country']
        if not name:
            return Country.browse()
        return Country.search(
            ['|', ('name', '=ilike', name), ('code', '=ilike', name)], limit=1)

    @api.model
    def _find_state(self, name, country):
        State = self.env['res.country.state']
        if not name:
            return State.browse()
        domain = ['|', ('name', '=ilike', name), ('code', '=ilike', name)]
        if country:
            domain = ['&', ('country_id', '=', country.id)] + domain
        return State.search(domain, limit=1)

    # ------------------------------------------------------------------
    # Persisting / presenting results
    # ------------------------------------------------------------------
    def _store_results(self, results, col_index):
        Line = self.env['contact.upload.wizard.line']
        Line.create([dict(line, wizard_id=self.id) for _row, line in results])

        self.write({
            'state': 'done',
            'total_rows': len(results),
            'account_created_count': sum(
                1 for _r, l in results if l['account_status'] == 'Created'),
            'account_existing_count': sum(
                1 for _r, l in results
                if l['account_status'] in ('Already Exists', 'Reused (in file)')),
            'contact_created_count': sum(
                1 for _r, l in results if l['contact_status'] == 'Created'),
            'contact_skipped_count': sum(
                1 for _r, l in results if l['contact_status'].startswith('Skipped')),
            'error_count': sum(
                1 for _r, l in results if l['row_status'] == 'failed'),
            'result_file': self._build_result_file(results, col_index),
            'result_file_name': 'Contact Upload Result.xlsx',
        })

    def _build_result_file(self, results, col_index):
        """Write a re-uploadable .xlsx: the original input columns (canonical
        order) followed by the status columns, one row per processed line."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Result'

        headers = [label for _key, label in INPUT_KEYS] + [
            'Account Status', 'Contact Status', 'Row Status', 'Message']
        sheet.append(headers)
        header_fill = PatternFill('solid', fgColor='305496')
        for cell_obj in sheet[1]:
            cell_obj.font = Font(bold=True, color='FFFFFF')
            cell_obj.fill = header_fill
            cell_obj.alignment = Alignment(horizontal='center')

        fills = {
            'success': PatternFill('solid', fgColor='C6EFCE'),
            'skipped': PatternFill('solid', fgColor='FFEB9C'),
            'failed': PatternFill('solid', fgColor='FFC7CE'),
        }
        for raw_row, line in results:
            def cval(key):
                idx = col_index.get(key)
                if idx is None or idx >= len(raw_row):
                    return ''
                return self._clean(raw_row[idx])

            values = [cval(key) for key, _label in INPUT_KEYS]
            values += [
                line['account_status'],
                line['contact_status'],
                ROW_STATUS_LABEL.get(line['row_status'], line['row_status']),
                line['message'],
            ]
            sheet.append(values)
            fill = fills.get(line['row_status'])
            if fill:
                for cell_obj in sheet[sheet.max_row]:
                    cell_obj.fill = fill

        widths = [14, 12, 26, 14, 22, 18, 20, 22, 18, 16, 14, 12, 14, 14, 14, 18, 26, 12, 50]
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
        sheet.freeze_panes = 'A2'

        buffer = io.BytesIO()
        workbook.save(buffer)
        return base64.b64encode(buffer.getvalue())

    # ------------------------------------------------------------------
    # Reset for another upload
    # ------------------------------------------------------------------
    def action_reset(self):
        self.ensure_one()
        self.line_ids.unlink()
        self.write({
            'state': 'draft',
            'upload_file': False,
            'file_name': False,
            'result_file': False,
            'result_file_name': False,
            'total_rows': 0,
            'account_created_count': 0,
            'account_existing_count': 0,
            'contact_created_count': 0,
            'contact_skipped_count': 0,
            'error_count': 0,
        })
        return {
            'type': 'ir.actions.act_window',
            'name': _('Upload Contacts & Accounts'),
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }


class ContactUploadWizardLine(models.TransientModel):
    _name = 'contact.upload.wizard.line'
    _description = 'Upload Result Line'
    _order = 'row_number'

    wizard_id = fields.Many2one(
        'contact.upload.wizard', string='Wizard', ondelete='cascade')
    row_number = fields.Integer(string='Row')
    company_name = fields.Char(string='Account')
    website = fields.Char(string='Website')
    contact_name = fields.Char(string='Contact')
    email = fields.Char(string='Email')
    account_status = fields.Char(string='Account Status')
    contact_status = fields.Char(string='Contact Status')
    row_status = fields.Selection([
        ('success', 'Success'),
        ('skipped', 'Skipped'),
        ('failed', 'Failed'),
    ], string='Status')
    message = fields.Char(string='Message')
    company_id = fields.Many2one('res.partner', string='Account Record')
    partner_id = fields.Many2one('res.partner', string='Contact Record')
